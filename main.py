# 修改时间22-4-6
# 采用遗传算法进行日尺度内，小时尺度下水库调度计算
# 采用小时径流数据以及风光小时出力数据
# 采用geatpy遗传算法库，采用带精英保留的遗传算法求解
# 调度目标是日出力过程的波动性最小,即小时出力变化数据的方差最小
# 创建人:姬新洋

import time

import matplotlib.pyplot as plt  # 加载绘图模块
import geatpy as ea  # 加载geatpy库
import numpy as np  # 加载矩阵运算模块
import openpyxl


# 加载计算数据表格，并导入风光出力数据以及径流数据
# 导入上游水位边界条件数据，水电小时流量数据
wb = openpyxl.load_workbook('jisuanshuju1.xlsx')
sheet1 = wb.get_sheet_by_name('上游水位变化数据')
sheet2 = wb.get_sheet_by_name('入库流量（1h尺度）')
sheet3 = wb.get_sheet_by_name('风电出力数据（1h尺度）')
sheet4 = wb.get_sheet_by_name('光电出力数据（1h尺度）')

h1_bianjietiaojian = [0 for _ in range(365)]
for i in range(2, 367):
    h1_bianjietiaojian[i - 2] = sheet1.cell(row=i, column=1).value  # 导入上游水位边界数据

q1_rukushuju = [0 for _ in range(8760)]
for i in range(2, 8762):
    q1_rukushuju[i - 2] = sheet2.cell(row=i, column=2).value  # 导入入库流量年内小时尺度数据

n1_fengshuju = [0 for _ in range(8760)]
for i in range(2, 8762):
    n1_fengshuju[i - 2] = sheet3.cell(row=i, column=2).value  # 导入风出力年内小时尺度数据

n1_guangshuju = [0 for _ in range(8760)]
for i in range(2, 8762):
    n1_guangshuju[i - 2] = sheet4.cell(row=i, column=2).value  # 导入光出力年内小时尺度数据

# ====================================导入数据部分完成======================================

# 存入边界限定条件变量中
# 这里接入水电的边界条件应为365天的日水位变化
h_shuibianjie = [0 for _ in range(2)]  # 初始化边界条件列表
h_shuibianjie[0] = h1_bianjietiaojian[0]
h_shuibianjie[1] = h1_bianjietiaojian[1]


# 存入水电站小时流量数据数组中
# 取第一天的入库流量数据
q_shuiruku = np.zeros((1, 24))  # 初始化数组
for i in range(24):
    q_shuiruku[[0], [i]] = q1_rukushuju[i]

# 这里取第一天的风光出力数据
n_feng = np.zeros((1, 24))  # 初始化数据
for i in range(24):
    n_feng[[0], [i]] = n1_fengshuju[i]

n_guang = np.zeros((1, 24))  # 初始化数组
for i in range(24):
    n_guang[[0], [i]] = n1_guangshuju[i]


# 这里接入风光365天的小时出力数据
# 格式应为365行24列，风光出力的单位为兆瓦
# n_feng = np.random.randint(200, 500, size=(1, 24))  # 这里用随机数模拟数据用于测试
# n_guang = np.random.randint(200, 500, size=(1, 24))  # 这里用随机数模拟数据用于测试


# 根据上游库容计算上游水位
def h1_v(vx):
    h = (-8e-6) * float(vx) ** 4 + 0.0022 * float(vx) ** 3 - 0.2104 * float(vx) ** 2 + 9.5804 * float(vx) + 1659.1
    return h


# 定义根据上游水位计算上游库容函数 计算结果库容是亿立方米
def v1(hx):
    v = 3.227e-6 * (float(hx) ** 3) - 0.01535 * (float(hx) ** 2) + (24.32 * float(hx)) - 1.283 * (10 ** 4)
    return v


# 定义根据下泄流量计算下游水位的函数
def h2(qx):
    h = 7e-12 * float(qx) ** 3 - 2e-7 * float(qx) ** 2 + 0.0034 * float(qx) + 1633.6
    return h


# 定义出力计算函数（输入量仅为时段初末的上游水位，还有该时段内的入库流量）计算结果是兆瓦
def n12(h1x, h2x, qx):
    n = 0
    h_pingjun = (float(h1x) + float(h2x)) / 2
    q_xiayou = ((v1(float(h1x)) - v1(float(h2x))) * (10 ** 8) / (60 * 60)) + float(qx)
    if q_xiayou > 14100:  # 防止下泄流量过大导致拟合函数失真
        q_xiayou = 14100
    h_xiayou = h2(q_xiayou)
    if q_xiayou < 0:
        n = 0  #
    elif q_xiayou > 2024.4:
        n = 8.3 * (h_pingjun - h_xiayou) * 2024.4 / 1000
        if n < 0:  # 最小出力限制条件
            n = 0
        elif n > 3600:  # 最大出力限制条件(装机容量)
            n = 3600
    else:
        n = 8.3 * (h_pingjun - h_xiayou) * q_xiayou / 1000
        if n < 0:  # 最小出力限制条件
            n = 0
        elif n > 3600:  # 最大出力限制条件(装机容量)
            n = 3600
    return n


# 目标函数
def aim(Phen):  # 传入种群染色体矩阵解码后的基因表现型矩阵（水电水位数据矩阵）
    # 这里应该把输入的23位水位数据转化为24位的出力数据
    Chuli = np.zeros((NIND, 24))  # 初始化出力表现型矩阵

    for i in range(NIND):
        n_shuiwei = np.array(Phen[[i], :])
        Chuli[[i], [0]] = n12(h_shuibianjie[0], n_shuiwei[[0], [0]], q_shuiruku[[0], [0]])
        Chuli[[i], [23]] = n12(n_shuiwei[[0], [22]], h_shuibianjie[1], q_shuiruku[[0], [23]])
        for j in range(1, 23):
            Chuli[[i], [j]] = n12(n_shuiwei[[0],[j - 1]], n_shuiwei[[0], [j]], q_shuiruku[[0], [j]])

    x_std = np.arange(NIND).reshape(NIND, 1)
    for i in range(NIND):
        n_zong = np.array(Chuli[[i], :]) + np.array(n_feng) + np.array(n_guang)
        # 通道限制
        for j in range(24):
            if n_zong[[0], [j]] > 3600:
                n_zong[[0], [j]] = 3600

        x_std[i, 0] = np.std(n_zong, ddof=1)

    return np.array(x_std)


# 变量设置 这里变量范围的优化将会在以后进行修改
# 这里设置决策变量的变化范围（各个时间段水电出力的有效范围）水电最大出力3600兆瓦
x_bianjie = [0 for _ in range(23)]  # 初始化变量边界生成矩阵
x_bianjie[0] = h1_v(v1(h_shuibianjie[0]) +
                    ((q_shuiruku[[0], [0]]) * (60 * 60) / (10 ** 8))
                    )
for i in range(1, 23):
    x_bianjie[i] = h1_v(v1(x_bianjie[i - 1]) +
                        ((q_shuiruku[[0], [i]]) * (60 * 60) / (10 ** 8))
                        )

x1 = [1800, x_bianjie[0]]  # 第一个决策变量的范围（后面依次类推）
x2 = [1800, x_bianjie[1]]
x3 = [1800, x_bianjie[2]]
x4 = [1800, x_bianjie[3]]
x5 = [1800, x_bianjie[4]]
x6 = [1800, x_bianjie[5]]
x7 = [1800, x_bianjie[6]]
x8 = [1800, x_bianjie[7]]
x9 = [1800, x_bianjie[8]]
x10 = [1800, x_bianjie[9]]
x11 = [1800, x_bianjie[10]]
x12 = [1800, x_bianjie[11]]
x13 = [1800, x_bianjie[12]]
x14 = [1800, x_bianjie[13]]
x15 = [1800, x_bianjie[14]]
x16 = [1800, x_bianjie[15]]
x17 = [1800, x_bianjie[16]]
x18 = [1800, x_bianjie[17]]
x19 = [1800, x_bianjie[18]]
x20 = [1800, x_bianjie[19]]
x21 = [1800, x_bianjie[20]]
x22 = [1800, x_bianjie[21]]
x23 = [1800, x_bianjie[22]]  # 第二十三个水位决策变量的范围
# 这里设置决策变量的边界信息（包含或不包含）
b1 = [1, 1]  # 第一个决策变量边界，1表示包含范围的边界，0表示不包含（后面依次类推）
b2 = [1, 1]
b3 = [1, 1]
b4 = [1, 1]
b5 = [1, 1]
b6 = [1, 1]
b7 = [1, 1]
b8 = [1, 1]
b9 = [1, 1]
b10 = [1, 1]
b11 = [1, 1]
b12 = [1, 1]
b13 = [1, 1]
b14 = [1, 1]
b15 = [1, 1]
b16 = [1, 1]
b17 = [1, 1]
b18 = [1, 1]
b19 = [1, 1]
b20 = [1, 1]
b21 = [1, 1]
b22 = [1, 1]
b23 = [1, 1]

# 生成自变量的范围矩阵，使得第一行为所有决策变量的下界，第二行为上界
ranges = np.vstack([x1, x2, x3, x4, x5, x6, x7, x8, x9, x10, x11, x12, x13, x14, x15, x16, x17,
                    x18, x19, x20, x21, x22, x23]).T
# 生成自变量的边界矩阵
borders = np.vstack([b1, b2, b3, b4, b5, b6, b7, b8, b9, b10, b11, b12, b13, b14, b15, b16, b17,
                     b18, b19, b20, b21, b22, b23]).T
# 设置决策变量的类型，0表示连续，1表示离散（这里24个点位均为离散点位）
varTypes = np.array([1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1])

# 染色体编码设置
Encoding = 'BG'  # 'BG'表示采用二进制/格雷编码
# 决策变量的编码方式，1表示变量均使用格雷编码
codes = [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
# 决策变量的编码精度，表示解码后能表示的决策变量的精度可达到小数点后6位
precisions = [3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3]
scales = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]  # 0表示采用算术刻度，1表示采用对数刻度
# 调用函数创建译码矩阵
FieldD = ea.crtfld(Encoding, varTypes, ranges, borders, precisions, codes, scales)


# 遗传算法参数设置
NIND = 30  # 种群个体数目
MAXGEN = 3000  # 最大遗传代数
# 表示目标函数是最小化，元素为-1时则表示对应的目标函数是最大化
maxormins = np.array([1])
selectStyle = 'sus'  # 采用随机抽样选择
recStyle = 'xovdp'  # 采用两点交叉
mutStyle = 'mutbin'  # 采用二进制染色体的变异算子
Lind = int(np.sum(FieldD[0, :]))  # 计算染色体长度
pc = 0.9  # 交叉概率
pm = 1 / Lind  # 变异概率

# 定义目标函数值记录器
obj_trace = np.zeros((MAXGEN, 23))
# 定义染色体记录器，记录历代最优个体的染色体
var_trace = np.zeros((MAXGEN, Lind))


# 开始遗传算法进化
start_time = time.time()  # 开始计时
Chrom = ea.crtpc(Encoding, NIND, FieldD)  # 生成种群染色体矩阵
variable = ea.bs2ri(Chrom, FieldD)  # 对初始种群进行解码
ObjV = aim(variable)  # 计算初始种群个体的目标函数值
best_ind = np.argmin(ObjV)  # 计算当代最优个体的序号


# 开始进化
for gen in range(MAXGEN):
    FitnV = ea.ranking(maxormins * ObjV)  # 根据目标函数的大小分配适应度值
    SelCh = Chrom[ea.selecting(selectStyle, FitnV, NIND - 1), :]  # 选择
    SelCh = ea.recombin(recStyle, SelCh, pc)  # 重组
    SelCh = ea.mutate(mutStyle, Encoding, SelCh, pm)  # 变异
    # 把父代精英个体与子代的染色体进行合并，得到新一代种群
    Chrom = np.vstack([Chrom[best_ind, :], SelCh])
    Phen = ea.bs2ri(Chrom, FieldD)  # 对种群进行解码（二进制转十进制）
    ObjV = aim(Phen)  # 求种群个体的目标函数值
    # 记录
    best_ind = np.argmin(ObjV)  # 计算当代最优个体的序号
    obj_trace[gen, 0] = np.sum(ObjV) / NIND  # 记录当代种群的目标函数均值
    obj_trace[gen, 1] = ObjV[best_ind]  # 记录当代种群最优个体目标函数值
    var_trace[gen, :] = Chrom[best_ind, :]  # 记录当代种群最优个体的染色体
# 进化完成


# 输出结果处理
best_gen = np.argmin(obj_trace[:, [1]])  # 最优解的位置代号
variable = ea.bs2ri(var_trace[[best_gen], :], FieldD)  # 解码得到表现型矩阵

# 计算表现型矩阵出力表达量
n_variable = np.zeros((1, 24))
n_variable[[0], [0]] = n12(h_shuibianjie[0], variable[[0], [0]], q_shuiruku[[0], [0]])
n_variable[[0], [23]] = n12(variable[[0], [22]], h_shuibianjie[1], q_shuiruku[[0], [23]])
for j in range(1, 23):
    n_variable[[0], [j]] = n12(variable[[0], [j - 1]], variable[[0], [j]], q_shuiruku[[0], [j]])


# 计算运行时间
end_time = time.time()  # 结束计时
time_c = end_time - start_time  # 计算计时时间
print('本次计算用时:' + str(time_c) + 's')

# 图像绘制部分需要进行修改
# 进行图像绘制
ea.trcplot(obj_trace, [['种群个体平均目标函数值', '种群最优个体目标函数值']])  # 绘制种群最优个体函数值进化图像
# 绘制目标结果示意曲线
n_niheshui = n_variable[[0], :]  # 取出当前最优个体的数据
n_nihezong = np.array(n_niheshui) + np.array(n_feng) + np.array(n_guang)
x_zuobiao = np.array([1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24]).reshape(24)
# 转换数据格式
y1_tu = n_nihezong.T
y2_tu = n_niheshui.T
y3_tu = n_feng.T
y4_tu = n_guang.T
plt.figure(1, figsize=(16, 8))  # 设置图形的像素大小
plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来显示正常中文标签
plt.plot(x_zuobiao, y1_tu, label='总出力', color='black', linewidth=2)
plt.plot(x_zuobiao, y2_tu, label='水电机组计划出力', color='blue', linewidth=2)
plt.plot(x_zuobiao, y3_tu, label='风电出力', color='red', linewidth=2)
plt.plot(x_zuobiao, y4_tu, label='光伏出力', color='yellow', linewidth=2)
plt.xlabel('小时时刻')
plt.ylabel('出力（兆瓦）')
plt.title('日出力曲线')
plt.legend()

plt.show()  # 显示图形
